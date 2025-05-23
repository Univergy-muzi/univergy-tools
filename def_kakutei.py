def analyze_kakutei_xml(xml_file_list):
  import xmltodict
  import pandas as pd
  import datetime as dt
  import xml.etree.ElementTree as ET
  from openpyxl import Workbook
  from io import BytesIO

  wb = Workbook()
  ws = wb.active
  ws.title = 'DATA'
  for i in range(2):
      ws.merge_cells(start_row=1,
                     start_column=i + 1,
                     end_row=2,
                     end_column=i + 1)
  ws.cell(row=1, column=1, value='DATE')
  ws.cell(row=1, column=2, value='TIME')
  ws.cell(row=1, column=3, value='TOTAL')

  ws.sheet_properties.tabColor = "73D790"

  #####################################################################################################################################

  field = ['NUMBER', 'NAME', 'DATE', 'TIME', 'VALUE']

  for x in xml_file_list:
      df = pd.DataFrame(columns=field)
      print(x)

      tree = ET.parse(x)
      xml_data = tree.getroot()
      xmlstr = ET.tostring(xml_data, encoding='utf-8', method='xml')

      result = xmltodict.parse(xmlstr)
      result_list = result['SBD-MSG']['JPMGRP']['JPTRM']['JPM00010'][
          'JPMR00010']
      try:
          if result_list['JP06405'] == "1":
              continue
          check_list = result_list['JPM00013']['JPMR00013']
          tmp_list = []
          final_list = []
          range_check = len(check_list)
          for y in range(range_check):
              ref_list = check_list[y]
              ref2_list = ref_list['JPM00014']['JPMR00014']
              len_check = len(ref2_list)
              for a in range(len_check):
                  tmp_list.append(result_list['JP06400'])
                  tmp_list.append(result_list['JP06120'])
                  tmp_list.append(ref_list['JP06423'])
                  tmp_list.append(ref2_list[a]['JP06219'])
                  tmp_list.append(ref2_list[a]['JP06424'])
                  final_list.append(tmp_list)
                  tmp_list = []

              _df = pd.DataFrame(final_list, columns=field)
              _df = _df.fillna(0)
              df = pd.concat([df, _df])
              final_list = []
      except:
          range_check = len(result_list)
          for y in range(range_check):
              if result_list[y]['JP06405'] == "1":
                  continue
              check_list = result_list[y]['JPM00013']['JPMR00013']
              tmp_list = []
              final_list = []
              range_check2 = len(check_list)
              for z in range(range_check2):
                  ref_list = check_list[z]
                  ref2_list = ref_list['JPM00014']['JPMR00014']
                  len_check = len(ref2_list)
                  for a in range(len_check):
                      tmp_list.append(result_list[y]['JP06400'])
                      tmp_list.append(result_list[y]['JP06120'])
                      tmp_list.append(ref_list['JP06423'])
                      tmp_list.append(ref2_list[a]['JP06219'])
                      tmp_list.append(ref2_list[a]['JP06424'])
                      final_list.append(tmp_list)
                      tmp_list = []

                  _df = pd.DataFrame(final_list, columns=field)
                  _df = _df.fillna(0)
                  df = pd.concat([df, _df])
                  final_list = []

#####################################################################################################################################

      df = df.sort_values(by=['DATE', 'TIME'])

      dates = df.loc[:, 'DATE'].tolist()
      dates = list(dict.fromkeys(dates))

      target_date_list = []
      if ws.max_row > 3:
          for i in range(int((ws.max_row - 2) / 48)):
              tmp_d = ws.cell(row=3 + (i * 48), column=1).value
              tmp_d = dt.date(tmp_d.year, tmp_d.month, tmp_d.day)
              target_date_list.append(tmp_d)

      for s in dates:
          t_date = dt.date(int(s[:4]), int(s[4:6]), int(s[6:]))
          if t_date not in target_date_list:
              end_row = ws.max_row
              for i in range(48):
                  ws.cell(row=end_row + i + 1, column=1, value=t_date)
                  ws.cell(row=end_row + i + 1, column=2, value=i + 1)

#####################################################################################################################################

      nos = df.loc[:, 'NUMBER'].tolist()
      nos = list(dict.fromkeys(nos))

      target_no_list = []
      if ws.max_column > 3:
          for i in range(ws.max_column - 3):
              target_no_list.append(ws.cell(row=2, column=i + 3).value)

      for s in nos:
          target_date = df.loc[df.NUMBER == s, 'DATE'].tolist()[0]
          target_date = dt.date(int(target_date[:4]), int(target_date[4:6]),
                                int(target_date[6:]))
          for i in range(int((ws.max_row - 2) / 48)):
              tmp_d = ws.cell(row=3 + (i * 48), column=1).value
              tmp_d = dt.date(tmp_d.year, tmp_d.month, tmp_d.day)
              if tmp_d == target_date:
                  target_row = 3 + (i * 48)
                  break

          target_column = 0
          if s not in target_no_list:
              t_name = df.loc[df.NUMBER == s, 'NAME'].tolist()[0]
              ws.insert_cols(ws.max_column)
              ws.cell(row=1, column=ws.max_column - 1, value=t_name)
              ws.cell(row=2, column=ws.max_column - 1, value=s)
              target_column = ws.max_column - 1

          if target_column == 0:
              for i in range(ws.max_column - 3):
                  if ws.cell(row=2, column=i + 3).value == s:
                      target_column = i + 3
                      break

          t_value_list = df.loc[df.NUMBER == s, 'VALUE'].tolist()

          for i in range(len(t_value_list)):
              ws.cell(row=target_row + i,
                      column=target_column,
                      value=float(t_value_list[i]))

#####################################################################################################################################

  if 64 + ws.max_column > 116:
      ws.cell(row=3,
              column=ws.max_column,
              value=f'=ROUND(SUM(C3:B{chr(64+ws.max_column-53)}3),1)')
  elif 64 + ws.max_column > 90:
      ws.cell(row=3,
              column=ws.max_column,
              value=f'=ROUND(SUM(C3:A{chr(64+ws.max_column-27)}3),1)')
  else:
      ws.cell(row=3,
              column=ws.max_column,
              value=f'=ROUND(SUM(C3:{chr(64+ws.max_column-1)}3),1)')


#####################################################################################################################################

  now = dt.datetime.now()
  now = now.strftime(r'%y-%m-%d_%H%M')

  excel_stream = BytesIO()
  wb.save(excel_stream)
  wb.close
  excel_stream.seek(0)

  return excel_stream
